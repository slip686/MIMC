import ast
from Custom_Widgets.Widgets import QMainWindow
from PySide6.QtWidgets import QApplication, QFileDialog, QGridLayout
from CustomWidgets import *
from UIwindow import Ui_MainWindow
from core import *
from PySide6.QtGui import Qt, QCursor, QAction, QPainter, QPixmap, QTextCursor, QBrush, QColor
from PySide6.QtWidgets import QTreeWidgetItemIterator
from openpyxl import load_workbook
from core import get_advice
from Dialogs import *
import schedule
import time
from threading import Thread


# this shit is just for correct SQL strings
class str2(str):
    def __repr__(self):
        return ''.join(('"', super().__repr__()[1:-1], '"'))


class MainWindow(QMainWindow):

    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.show()
        self.ui.mainMenuStack.setCurrentIndex(0)
        self.ui.regStackedWidget.setCurrentWidget(self.ui.loginPage)
        self.session = user_connection(None, None, self.ui.statusLabel2, main_window_object=self)
        self.logged_user_data = None
        self.logged_user = User()
        self.projects_ids_and_roles_list = []
        self.current_project_id = None
        self.structure_table_names = None
        self.current_project_data_dict = None
        self.current_project_docs_dicts_list = None
        self.current_project_users_data = None
        self.received_notifications = None
        self.project_widget_list = []
        self.ui.designDocsStructureTreeWidget.patterns_list = []
        self.ui.constructionDocsStructureTreeWidget.patterns_list = []
        self.ui.initialPermitDocsStructureTreeWidget.patterns_list = []
        self.ui.interfaceBodyStackedWidget.setCurrentIndex(1)
        self.ui.homeBtn.clicked.connect(lambda: self.ui.interfaceBodyStackedWidget.slideInIdx(1))
        self.ui.flowlayout = FlowLayout(parent=self.ui.widget_4, margin=40, spacing=25)
        self.ui.scrollArea.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.ui.homeBtn.hide()
        self.ui.leftSideMenuBtn.hide()
        self.ui.leftSidePopUpMenu.hide()
        self.document_filepath = None
        self.ui.backToStructureBtn.hide()
        self.ui.backToStructureBtn_2.hide()
        self.ui.backToStructureBtn_3.hide()
        self.current_design_docs_folder = None
        self.current_design_docs_folder_path = 'All Documents'
        self.current_construction_docs_folder = None
        self.current_construction_docs_folder_path = 'All Documents'
        self.current_init_permission_docs_folder = None
        self.current_init_permission_docs_folder_path = 'All Documents'
        self.users_data_from_db = None

        self.ui.designDocsTableWidget.main_table = True
        self.ui.constructionDocsTableWidget.main_table = True
        self.ui.initDocsTableWidget.main_table = True

        self.message_receiver = None
        self.message_receiving_thread = None

        self.ui.horizontalLayout_57.removeWidget(self.ui.NotificationsMenu)
        self.ui.NotificationsMenu.setParent(self.ui.interfaceBodySubContainer)
        self.ui.NotificationsMenu.hide()
        self.ui.NotificationsMenu.session_object = self.session
        self.ui.NotificationsMenu.main_window = self

        if get_platform() == 'win':
            self.ui.tabWidget.setStyleSheet(u"#tabWidget::pane {background-color: rgb(136,136,136)}\n"
                                            "#tabWidget ::tab::!selected {background-color: rgb(136, 136, 136)}\n"
                                            "#tabWidget ::tab::selected {border: 2px solid rgb(67, 67, 67)}\n"
                                            "#tabWidget ::tab::selected {border-radius: 6px }\n"
                                            "#tabWidget ::tab::selected {padding: 3px }\n"
                                            "\n"
                                            "\n"
                                            "QSplitter::handle:horizontal {border-radius: 3px; background-color: rgb(165, 165, 165);}\n"
                                            ""
                                            u'QTabWidget::tab-bar {alignment: center;}'
                                            u'QTabBar {color: white}')

        ############################################################################################
        # getting advice
        ############################################################################################
        def show_advice():
            self.ui.statusLabel.setText(get_advice())

        self.stop = False

        def run(stop_thread):
            schedule.every(10).seconds.do(show_advice)
            while True:
                schedule.run_pending()
                time.sleep(1)
                if stop_thread():
                    break

        thread = Thread(target=run, args=(lambda: self.stop,), name='show_advice')
        thread.start()

        ############################################################################################
        # DOCS STRUCTURE LISTS
        self.design_docs_structure_list = None
        self.construction_docs_structure_list = None
        self.initial_permit_docs_structure_list = None
        self.design_docs_tree_rows = None
        self.construction_docs_tree_rows = None
        self.initial_permit_docs_tree_rows = None
        ############################################################################################

        self.ui.tabWidget.currentChanged.connect(lambda: tab_changing())
        self.ui.stackedWidget_3.setCurrentIndex(0)
        self.current_item_index = None

        def tab_changing():
            if self.ui.tabWidget.currentIndex() == 0:
                self.ui.stackedWidget_3.slideInIdx(0)
            if self.ui.tabWidget.currentIndex() == 1:
                self.ui.stackedWidget_3.slideInIdx(1)
            if self.ui.tabWidget.currentIndex() == 2:
                self.ui.stackedWidget_3.slideInIdx(2)

        def clear_widgets():
            for i in reversed(range(self.ui.flowlayout.count())):
                self.ui.flowlayout.itemAt(i).widget().deleteLater()

        def get_project_cards():
            clear_widgets()
            self.project_widget_list.clear()
            user_projects = self.session.api.get_projects_ids_and_roles(self.logged_user.user_id)
            if user_projects:
                for key in user_projects.get("content"):
                    project_data = self.session.api.get_project(key)
                    if project_data.get("code") == 200:
                        project_data_dict = project_data.get("content")
                        project_owner = self.session.api.get_user_data(project_data_dict.get("owner_id"))
                        self.project_widget = ProjectCard(data_dict=project_data_dict,
                                                          main_window=self,
                                                          role=user_projects.get(key))
                        self.project_widget.project_id = project_data_dict['id']
                        self.project_widget.projectName.setText(project_data_dict['project_name'])
                        self.project_widget.Owner.setText(f'Owner: {project_owner.get("content").get("first_name")} '
                                                          f'{project_owner.get("content").get("last_name")}')
                        self.project_widget.ConstrucionStatus.setText(project_data_dict['status'])
                        self.project_widget.Time.setText(project_data_dict['time_limits'])
                        self.project_widget.project_picture = project_data_dict['picture']
                        self.project_widget_list.append(self.project_widget)

                for i in self.project_widget_list:
                    self.ui.flowlayout.addWidget(i)

                    def set_picture(label, stop):
                        while True:
                            time.sleep(1)
                            if stop():
                                break
                            if self.session.connection_success:
                                picture_bytes = self.session.download_process(
                                    repo_id=i.project_data_dict['repo_id'],
                                    file_address=f'/{i.project_data_dict["picture"]}',
                                    bytes_format=True)
                                if picture_bytes:
                                    picture = QPixmap()
                                    picture.loadFromData(picture_bytes)
                                    rounded = QPixmap(picture.size())
                                    rounded.fill(QColor("transparent"))
                                    painter = QPainter(rounded)
                                    painter.setRenderHint(QPainter.Antialiasing)
                                    painter.setBrush(QBrush(picture))
                                    painter.setPen(Qt.NoPen)
                                    painter.drawRoundedRect(picture.rect(), 15, 15)
                                    painter.end()
                                    try:
                                        label.setPixmap(rounded)
                                    except RuntimeError:
                                        pass
                                    break

                    set_picture_thread = Thread(target=set_picture, args=(i.label, lambda: self.stop),
                                                name=f'picture thread {i}')
                    set_picture_thread.start()

                self.ui.flowlayout.parent().parent().parent().parent().set_scroll_bar_parameters()

            else:
                self.ui.statusLabel.setText('No projects yet')

        def start_session(email=None, password=None, remember=None):

            self.session.email = email
            self.session.password = password
            self.session.remember = remember
            connection = self.session.session()

            if connection == 'connected':
                self.ui.mainMenuStack.slideInIdx(1)
                self.ui.interfaceBodyStackedWidget.setCurrentIndex(1)
                self.ui.emailEntering.setText('')
                self.ui.passEntering.setText('')
                self.message_receiver = NotificationReceiver(mq_connection_object=self.session.conn_broker,
                                                             session_object=self.session)
                self.message_receiving_thread = QThread()
                self.message_receiver.moveToThread(self.message_receiving_thread)
                self.message_receiving_thread.started.connect(self.message_receiver.start_broker_loop)
                self.message_receiver.got_message.connect(
                    lambda: self.add_notification(ntfcn_dict=self.message_receiver.message))
                for message in self.session.notifications:
                    self.ui.NotificationsMenu.insert_notification(ntfcn_dict=message)
                self.message_receiving_thread.start()

                self.logged_user_data = self.session.user_data
                self.logged_user.user_id = self.logged_user_data['user_id']
                self.logged_user.user_email = self.logged_user_data['email']
                self.logged_user.first_name = self.logged_user_data['first_name']
                self.logged_user.last_name = self.logged_user_data['last_name']
                self.logged_user.company_name = self.logged_user_data['company_name']
                self.logged_user.TIN = self.logged_user_data['tin']
                self.logged_user.notification_channel = self.logged_user_data['ntfcn_channel']
                self.ui.companyInfoBtn.setText(self.logged_user.company_name)

                get_project_cards()
                return 'connected'

            if connection is None:
                return
            else:
                self.ui.label_3.setText(connection.capitalize())
                return connection

        start_session()

        self.new_user = Reg_data(None, None)
        self.json_process = Json_process(None, None)
        self.exist_user = Reg_data(None, None)

        ########################################################################################################
        # _____________________________REGISTRATION, RECOVER, LOGIN PROCESSES___________________________________
        ########################################################################################################

        def sending_key():
            self.new_user.Email = self.ui.emailRegEntering.text()
            key = get_key(self.new_user.Email)
            self.new_user.Hash_key = hash(key)
            sending_process = Email_reg_sending(self.new_user.Email, key)
            if self.ui.termsAcception.checkState().value:
                response = self.session.api.check_email(self.new_user.Email)
                if response.get("code") == 404:
                    sending_process.start()
                    if not sending_process.correct_email:
                        self.ui.infoLabel_3.setText('Incorrect e-mail address')
                    elif not sending_process.successfully_sent:
                        self.ui.infoLabel_3.setText('No internet connection')
                    else:
                        self.ui.infoLabel_3.setText('')
                        self.ui.regStackedWidget.slideInIdx(2)
                        self.ui.infoLabel_4.setText('')
                if response.get("code") == 200:
                    self.ui.infoLabel_3.setText('Email already exists')
                if response.get("code") == 400:
                    self.ui.infoLabel_3.setText('No email entered')
            else:
                self.ui.infoLabel_3.setText('You should accept terms')

        def sending_recover_key():
            self.exist_user.Email = self.ui.emailRegEntering_2.text()
            key = get_key(self.exist_user.Email)
            self.exist_user.Hash_key = hash(key)
            sending_process = Email_recover_sending(self.exist_user.Email, key)
            sending_process.start()
            if not sending_process.correct_email:
                self.ui.infoLabel_7.setText('Incorrect e-mail address')
            elif not sending_process.successfully_sent:
                self.ui.infoLabel_7.setText('No internet connection')
            else:
                self.ui.regStackedWidget.slideInIdx(7)

        def key_accepting():
            self.ui.infoLabel_4.setText('')
            if hash(self.ui.keyEntering.text()) == self.new_user.Hash_key:
                self.ui.infoLabel_4.setText('Key accepted')
                self.ui.regStackedWidget.slideInIdx(3)
            else:
                self.ui.infoLabel_4.setText('Wrong key')

        def key_recover_accepting():
            self.ui.infoLabel_9.setText('')
            if hash(self.ui.keyEntering_2.text()) == self.exist_user.Hash_key:
                self.ui.infoLabel_9.setText('Key accepted')
                self.ui.regStackedWidget.slideInIdx(8)
            else:
                self.ui.infoLabel_9.setText('Wrong key')

        def new_user_password_entering():
            self.ui.infoLabel_5.setText('')
            self.new_user.password = self.ui.passEntering_2.text()
            if self.ui.passRepeatEntering.text() == self.new_user.password:
                self.ui.regStackedWidget.slideInIdx(4)
            else:
                self.ui.infoLabel_5.setText('Passwords are mismatch, try again')

        def user_password_recover():
            self.ui.infoLabel_10.setText('')
            self.exist_user.password = self.ui.passEntering_3.text()
            self.recover_process = push_recover_user_data(self.exist_user.Email,
                                                          self.exist_user.password)
            if self.ui.passRepeatEntering_2.text() == self.exist_user.password:
                self.recover_process.start()
                if self.recover_process.successful_insertion == 1:
                    exist_user_data_clear()
                    self.ui.regStackedWidget.slideInIdx(9)
                else:
                    self.ui.infoLabel_10.setText('Something went wrong')
            else:
                self.ui.infoLabel_10.setText('Passwords are mismatch, try again')

        def new_user_data_clear():
            self.new_user = Reg_data(None, None)
            self.new_user.name = None
            self.new_user.password = None
            self.new_user.last_name = None
            self.new_user.company_name = None
            self.new_user.job_title = None
            self.ui.emailRegEntering.setText('')
            self.ui.passEntering_2.setText('')
            self.ui.passRepeatEntering.setText('')
            self.ui.keyEntering.setText('')
            self.ui.nameEntering.setText('')
            self.ui.lastNameEntering.setText('')
            self.ui.companyTIN.setText('')
            self.ui.infoLabel_3.setText('')
            self.ui.termsAcception.setChecked(False)

        def new_project_data_begin_clear():
            self.ui.lineEdit_4.setText('')
            self.ui.lineEdit_3.setText('')
            self.ui.lineEdit_2.setText('')
            self.ui.lineEdit.setText('')

        def new_project_data_end_clear():
            self.ui.comboBox_4.clear()
            self.ui.comboBox_5.clear()
            self.ui.comboBox_6.clear()
            self.ui.comboBox_3.clear()
            self.ui.comboBox_7.clear()
            self.ui.comboBox_8.clear()
            self.ui.comboBox_9.clear()
            self.ui.comboBox_2.clear()

        def exist_user_data_clear():
            self.exist_user = Reg_data(None, None)
            self.exist_user.password = None
            self.ui.emailRegEntering_2.setText('')
            self.ui.passEntering_3.setText('')
            self.ui.passRepeatEntering_2.setText('')
            self.ui.keyEntering_2.setText('')
            self.ui.infoLabel_7.setText('')
            self.ui.infoLabel_9.setText('')

        def create_user():
            self.ui.infoLabel_6.setText('')
            self.new_user.name = self.ui.nameEntering.text()
            self.new_user.last_name = self.ui.lastNameEntering.text()
            self.new_user.TIN = self.ui.companyTIN.text()
            if self.new_user.name and self.new_user.last_name and self.new_user.TIN:
                if len(self.new_user.TIN) == 10:
                    company = get_company(self.new_user.TIN)
                    if company.get('status') == 'OK':
                        self.new_user.company_name = company.get('name')
                        response = self.session.api.create_user(self.new_user.as_dict)

                        if response.get("code") == 200:
                            new_user_data_clear()
                            self.ui.regStackedWidget.slideInIdx(5)
                        else:
                            self.ui.infoLabel_6.setText('Fault while writing to database')
                    else:
                        self.ui.infoLabel_6.setText('Fault while getting company data')
                else:
                    self.ui.infoLabel_6.setText('Wrong format TIN')
                    self.ui.companyTIN.setStyleSheet(u'background-color: rgb(184, 184, 184); color : black;'
                                                     u'border: 2px solid rgb(93, 0, 23)')
            else:
                if not self.new_user.name:
                    self.ui.nameEntering.setStyleSheet(u'background-color: rgb(184, 184, 184); color : black;'
                                                       u'border: 2px solid rgb(93, 0, 23)')
                if not self.new_user.last_name:
                    self.ui.lastNameEntering.setStyleSheet(u'background-color: rgb(184, 184, 184); color : black;'
                                                           u'border: 2px solid rgb(93, 0, 23)')
                if not self.new_user.TIN:
                    self.ui.companyTIN.setStyleSheet(u'background-color: rgb(184, 184, 184); color : black;'
                                                     u'border: 2px solid rgb(93, 0, 23)')

        def login_with_saving_option():
            if self.ui.rememberCheckBox.isChecked():
                start_session(email=self.ui.emailEntering.text(),
                              password=self.ui.passEntering.text(),
                              remember=True)
            else:
                start_session(email=self.ui.emailEntering.text(),
                              password=self.ui.passEntering.text())

        def highlight_unfilled(field, info_label=None, text=None):
            if isinstance(field, QComboBox):
                if not field.currentText():
                    field.setStyleSheet(
                        u"background-color: rgb(184, 184, 184); color : black; border-style: "
                        u"solid; border-width: 2px; border-color: rgb(93, 0, 23)")
                    if info_label:
                        info_label.setText(text)
                    return False
                else:
                    return True
            if isinstance(field, (QLineEdit, CQLineEdit)):
                if not field.text():
                    field.setStyleSheet(
                        u"background-color: rgb(184, 184, 184); color : black; border-style: "
                        u"solid; border-width: 2px; border-color: rgb(93, 0, 23)")
                    if info_label:
                        info_label.setText(text)
                    return False
                else:
                    return True

        def new_project_adding_process_page1():
            lineEdit_list = [self.ui.lineEdit_4, self.ui.lineEdit_3, self.ui.lineEdit_2, self.ui.lineEdit]
            lines_check_successful = False
            for line in lineEdit_list:
                lines_check_successful = highlight_unfilled(line, info_label=self.ui.label_10,
                                                            text='Unfilled parameter')

            if lines_check_successful:
                if self.session.api.project_exist(self.ui.lineEdit_3.text()).get("code") == 404:
                    self.ui.newProjectStackedWidget.slideInIdx(1)
                    self.ui.label_10.setText('')
                    self.users_data_from_db = self.session.api.users.get('content').get('users_list')
                    company_set = {item['company_name'] for item in self.users_data_from_db}
                    for i in company_set:
                        self.ui.comboBox_4.addItem(i)
                        self.ui.comboBox_5.addItem(i)
                        self.ui.comboBox_6.addItem(i)
                        self.ui.comboBox_3.addItem(i)
                    for num in range(self.ui.comboBox_4.count()):
                        if self.ui.comboBox_4.itemText(num) == self.logged_user.company_name:
                            self.ui.comboBox_4.setCurrentIndex(num)
                            for num_ in range(self.ui.comboBox_7.count()):
                                if self.ui.comboBox_7.itemText(num_) == f'{self.logged_user.last_name} ' \
                                                                        f'{self.logged_user.last_name}':
                                    self.ui.comboBox_7.setCurrentIndex(num_)
                else:
                    self.ui.label_10.setText('*Project name already exists')

        def set_company_users_list(company_name, combo_box):
            combo_box.clear()
            for item in self.users_data_from_db:
                if item['company_name'] == company_name:
                    combo_box.addItem(f'{item["last_name"]} {item["first_name"]}', userData={'id': item['user_id'],
                                                                                             'email': item['email']})

        def new_project_adding_process_page2():
            combo_boxes = [self.ui.comboBox_2, self.ui.comboBox_3, self.ui.comboBox_4, self.ui.comboBox_5,
                           self.ui.comboBox_6, self.ui.comboBox_7, self.ui.comboBox_8, self.ui.comboBox_9]
            combo_box_check_successful = False
            for box in combo_boxes:
                combo_box_check_successful = highlight_unfilled(box)
            if combo_box_check_successful:
                new_project = Project(picture=self.ui.lineEdit_4.text(), name=self.ui.lineEdit_3.text(),
                                      owner_id=self.logged_user.user_id, address=self.ui.lineEdit_2.text(),
                                      time=self.ui.lineEdit.text(), status=self.ui.comboBox.currentText(),
                                      user_connection_object=self.session)

                self.ui.statusLabel2.setText('Creating folders')
                new_project.add_folders_thread.start()
                new_project.add_folders_thread.join()
                users_emails_list = [self.ui.comboBox_7.itemData(self.ui.comboBox_7.currentIndex())['email'],
                                     self.ui.comboBox_8.itemData(self.ui.comboBox_8.currentIndex())['email'],
                                     self.ui.comboBox_2.itemData(self.ui.comboBox_2.currentIndex())['email'],
                                     self.ui.comboBox_9.itemData(self.ui.comboBox_9.currentIndex())['email']]

                new_project.add_users_to_stash_group(users_emails_list)

                data = {"project_data":
                            {"picture": new_project.picture.split('/')[-1],
                             "project_name": new_project.name,
                             "owner_id": new_project.owner_id,
                             "address": new_project.address,
                             "time_limits": new_project.time,
                             "status": new_project.status,
                             "repo_id": new_project.repo_id},
                        "users":
                            {
                                "user1":
                                    {
                                        "user_id": self.ui.comboBox_7.itemData(self.ui.comboBox_7.currentIndex())['id'],
                                        "job_title": "chief_engineer"
                                    },
                                "user2":
                                    {
                                        "user_id": self.ui.comboBox_8.itemData(self.ui.comboBox_8.currentIndex())['id'],
                                        "job_title": "contractor"
                                    },
                                "user3":
                                    {
                                        "user_id": self.ui.comboBox_2.itemData(self.ui.comboBox_2.currentIndex())['id'],
                                        "job_title": "designer"
                                    },
                                "user4":
                                    {
                                        "user_id": self.ui.comboBox_9.itemData(self.ui.comboBox_9.currentIndex())['id'],
                                        "job_title": "technical_client"
                                    }
                            }
                        }

                self.session.api.create_project(data)

                #######################################################################################

                self.ui.interfaceBodyStackedWidget.slideInIdx(1)
                new_project_data_begin_clear()
                new_project_data_end_clear()
                get_project_cards()

        def get_picture_filepath():
            self.dialogWindow = QFileDialog()
            filename = self.dialogWindow.getOpenFileName(None, "Select picture", "",
                                                         "Images (*.png *.tiff *.jpg *.jpeg)")
            self.ui.lineEdit_4.setText(filename[0])

        #######################################################################################################
        # __________________________________ DOCUMENTS STRUCTURE FUNCTIONS ____________________________________#
        #######################################################################################################

        def get_tree_widget_rows_contents_list(tree):
            rows_list = []
            iterator = QTreeWidgetItemIterator(tree)
            while iterator.value():
                row = iterator.value()
                if hasattr(row, 'place_id_list'):
                    rows_list.append([row.place_id_list, row.folder_name])
                iterator += 1
            return rows_list

        def get_rows_from_tree_widget():
            for tree in self.ui.stackedWidget_3.findChildren(QTreeWidget):
                if tree == self.ui.designDocsStructureTreeWidget:
                    self.design_docs_tree_rows = get_tree_widget_rows_contents_list(tree)
                if tree == self.ui.constructionDocsStructureTreeWidget:
                    self.construction_docs_tree_rows = get_tree_widget_rows_contents_list(tree)
                if tree == self.ui.initialPermitDocsStructureTreeWidget:
                    self.initial_permit_docs_tree_rows = get_tree_widget_rows_contents_list(tree)

        def edit_structure():
            for tree in self.ui.stackedWidget_3.findChildren(QTreeWidget):
                page = tree.parent().parent().parent().parent()
                current_page = self.ui.stackedWidget_3.currentWidget()
                if current_page == page:
                    self.iterator = QTreeWidgetItemIterator(tree)
                    while self.iterator.value():
                        self.item = self.iterator.value()
                        if hasattr(self.item, 'frame'):
                            self.item.frame.show()
                            self.item.fixRowElementHeightWhileEdit()
                            self.item.lineEdit.setStyleSheet(u"background-color: transparent")

                        self.iterator += 1

        def cancel_edit_structure():
            tree = None

            if self.ui.stackedWidget_3.currentIndex() == 0:
                tree = self.ui.designDocsStructureTreeWidget
            if self.ui.stackedWidget_3.currentIndex() == 1:
                tree = self.ui.constructionDocsStructureTreeWidget
            if self.ui.stackedWidget_3.currentIndex() == 2:
                tree = self.ui.initialPermitDocsStructureTreeWidget

            items_to_remove = []
            iterator = QTreeWidgetItemIterator(tree)
            while iterator.value():
                item = iterator.value()
                if isinstance(item, (RowElement,)):
                    items_to_remove.append(item)
                iterator += 1
            for i in items_to_remove:
                i.remove()

            self.get_project_structure()

        def update_structure_list():
            structure_list = None
            if self.ui.stackedWidget_3.currentIndex() == 0:
                design_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                          'design').get("content")
                if design_docs_query_result:
                    self.design_docs_structure_list = []
                    for i in design_docs_query_result:
                        list_element = list(i)
                        del list_element[1]
                        del list_element[1]
                        self.design_docs_structure_list.append(list_element)
                    structure_list = self.design_docs_structure_list
            elif self.ui.stackedWidget_3.currentIndex() == 1:
                construction_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                                'construction').get("content")
                if construction_docs_query_result:
                    self.construction_docs_structure_list = []
                    for i in construction_docs_query_result:
                        list_element = list(i)
                        del list_element[1]
                        del list_element[1]
                        self.construction_docs_structure_list.append(list_element)
                    structure_list = self.construction_docs_structure_list
            elif self.ui.stackedWidget_3.currentIndex() == 2:
                initial_permit_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                                  'init_permit').get("content")
                if initial_permit_docs_query_result:
                    self.initial_permit_docs_structure_list = []
                    for i in initial_permit_docs_query_result:
                        list_element = list(i)
                        del list_element[1]
                        del list_element[1]
                        self.initial_permit_docs_structure_list.append(list_element)
                    structure_list = self.initial_permit_docs_structure_list
            return structure_list

        def end_edit_structure():
            structure_list = update_structure_list()
            tree = None
            edit_slider = None
            most_long_pattern = 0
            current_patterns_list = []
            patterns_to_add = []
            docs_type = None

            if self.ui.stackedWidget_3.currentIndex() == 0:
                tree = self.ui.designDocsStructureTreeWidget
                edit_slider = self.ui.stackedWidget_4
                docs_type = 'design'
            elif self.ui.stackedWidget_3.currentIndex() == 1:
                tree = self.ui.constructionDocsStructureTreeWidget
                edit_slider = self.ui.stackedWidget_5
                docs_type = 'construction'
            elif self.ui.stackedWidget_3.currentIndex() == 2:
                tree = self.ui.initialPermitDocsStructureTreeWidget
                edit_slider = self.ui.stackedWidget_6
                docs_type = 'init_permit'

            database_cells_list = []
            current_cells_list = []
            if structure_list:
                for i in structure_list:
                    for num, value in enumerate(i[2:], 1):
                        if value:
                            database_cells_list.append([i[0], num, value])

            ##############################################################################################
            #                               DOCUMENTS STRUCTURE RULES
            # 1. Getting items indexes (index here is list of tuples) after clicking end edit button
            # 2. First tuple in list means (number + folder name) of top-level folder
            # 3. Following tuples mean level of child
            # 4. Each index (list of tuples) is a pattern, which has a number in database table
            # 5. Each document has a pattern number for sorting ability in interface table
            # 6. Each index writes as table string in database
            ##############################################################################################

            ##############################################################################################
            # 1. Search renamed and deleted rows and updating database, then updating docs_structure_list
            ##############################################################################################

            self.iterator = QTreeWidgetItemIterator(tree)
            while self.iterator.value():
                self.item = self.iterator.value()
                if hasattr(self.item, 'duplicate'):
                    if self.item.duplicate or not self.item.folder_name:
                        return self.ui.statusLabel.setText('Wrong row name')
                if hasattr(self.item, 'place_id_list'):
                    self.item.lineEdit.setFocus = False
                    if self.item.place_id_list:
                        for i in self.item.place_id_list:
                            current_cells_list.append([i, self.item.folder_name])
                    if self.item.renamed:
                        place_id_list = ','.join([str(i) for i in self.item.place_id_list])
                        # if len(self.item.place_id_list) == 1:
                        #     place_id_list = str(self.item.place_id_list)
                        #     print(place_id_list)
                        # else:
                        #     place_id_list = tuple([str(i) for i in self.item.place_id_list])
                        parent = self.item.parent()
                        child = self.item
                        rows_objects_list = []
                        while parent:
                            rows_objects_list.append(child)
                            child = parent
                            parent = parent.parent()
                        else:
                            rows_objects_list.append(child)
                        enumerated = enumerate(reversed(rows_objects_list), 1)
                        rows_objects_list_with_columns_nums = []
                        for num, obj in enumerated:
                            rows_objects_list_with_columns_nums.append((str(num), obj))

                        data = {"column_num": rows_objects_list_with_columns_nums[-1][0],
                                "new_value": self.item.folder_name,
                                "place_id_list": place_id_list}

                        self.session.api.update_place(data)

                        self.item.renamed = False

                self.iterator += 1

            ########################################
            # Deleting cells if row has been deleted
            ########################################

            structure_list = update_structure_list()

            database_cells_list = []
            if structure_list:
                for i in structure_list:
                    for num, value in enumerate(i[1:], 1):
                        if value:
                            database_cells_list.append([i[0], num, value])

            for i in database_cells_list:
                if [i[0], i[-1]] not in current_cells_list:
                    data = {"place_id": i[0], "column_num": i[1]}
                    self.session.api.exclude_place(data)
                    # self.session.commit_query([exclude_cell(), (AsIs(f'"{docs_structure_tablename}"'),
                    #                                             AsIs(f'"{i[1]}"'), AsIs('null'), AsIs(i[0]),
                    #                                             docs_type)])

            structure_list = update_structure_list()

            ########################################
            # Deleting empty patterns from database
            ########################################

            if structure_list:
                for i in structure_list:
                    if not i[1]:
                        self.session.api.delete_place(i[0])
                        # self.session.commit_query(
                        #     [delete_pattern(), (AsIs(f'{docs_structure_tablename}'), i[0], docs_type)])

            ######################################################################################
            # 2. Prepare list of patterns to add to database
            ######################################################################################

            structure_list_without_place_ids = []
            if structure_list:
                for i in structure_list:
                    m = list(i)
                    k = []
                    for j in m[1:]:
                        if j:
                            k.append(j)
                    structure_list_without_place_ids.append(tuple(k))

            self.iterator2 = QTreeWidgetItemIterator(tree)
            while self.iterator2.value():
                self.item = self.iterator2.value()
                if hasattr(self.item, 'place_id_list'):
                    folder_names_list = []
                    parent = self.item.parent()
                    child = self.item
                    while parent:
                        folder_names_list.append(child.folder_name)
                        child = parent
                        parent = parent.parent()
                    else:
                        folder_names_list.append(child.folder_name)
                    self.item.item_index = folder_names_list
                    if not self.item.childCount():
                        folder_names_tuple = tuple(reversed(folder_names_list))
                        current_patterns_list.append(folder_names_tuple)
                        if folder_names_tuple not in structure_list_without_place_ids:
                            patterns_to_add.append(folder_names_tuple)

                if hasattr(self.item, 'frame'):
                    self.item.frame.hide()
                if hasattr(self.item, 'place_id_list'):
                    if len(self.item.item_index) > most_long_pattern:
                        most_long_pattern = len(self.item.item_index)
                self.iterator2 += 1

            #####################################################################################################
            # 3. Adding columns if need
            #####################################################################################################

            # 1. Get names of columns (as ints) in current documents' table except 'place_id' column
            # (column named by its order in database table)

            columns_names_list = list(self.session.api.get_structure_columns_names)
            # columns_names_list = list(
            #     self.session.select_query([columns_names(), (AsIs(docs_structure_tablename),)], fetchall=True))
            del columns_names_list[0:3]

            #####################################################################################################
            # Adding columns process
            #####################################################################################################

            if len(columns_names_list) < most_long_pattern:
                delta = most_long_pattern - len(columns_names_list)
                columns_names_to_add = []
                self.column_name_to_add = 0
                if columns_names_list:
                    for i in range(1, delta + 1):
                        self.column_name_to_add = int(list(columns_names_list[-1])[0]) + i
                        columns_names_to_add.append(self.column_name_to_add)
                else:
                    for i in range(1, delta + 1):
                        self.column_name_to_add += 1
                        columns_names_to_add.append(self.column_name_to_add)
                for i in columns_names_to_add:
                    self.session.api.add_column_to_structure(i)
                    # self.session.commit_query([add_column(), (AsIs(docs_structure_tablename), AsIs(i))])

            #####################################################################################################
            # 4. Adding new patterns to database
            #####################################################################################################

            for i in patterns_to_add:
                columns_list = []
                values_list = []
                for column, value in enumerate(i, 1):
                    if len(i) == 1:
                        columns_single = f'{str(column)}'
                        values_single = value

                        data = {"columns": columns_single, "values": values_single}
                        self.session.api.add_place(project_id=self.current_project_data_dict['id'],
                                                   doctype=docs_type,
                                                   data=data)
                        # self.session.commit_query(
                        #     [insert_pattern(), (AsIs(docs_structure_tablename), AsIs(columns_single),
                        #                         AsIs('doc_type'), AsIs(values_single), docs_type)])
                    else:
                        columns_list.append(column)
                        values_list.append(value)
                if columns_list and values_list:
                    sql_string_columns = []
                    sql_string_values = []
                    for j in columns_list:
                        sql_string_columns.append(f'{j}')
                    for k in values_list:
                        sql_string_values.append(f"{k}")

                    data = {"columns": ','.join(sql_string_columns), "values": '|'.join(sql_string_values)}
                    self.session.api.add_place(project_id=self.current_project_data_dict['id'],
                                               doctype=docs_type,
                                               data=data)

                    # self.session.commit_query([insert_pattern(), (AsIs(docs_structure_tablename),
                    #                                               AsIs(', '.join(sql_string_columns)), AsIs('doc_type'),
                    #                                               AsIs(', '.join(sql_string_values)), docs_type)])

            #####################################################################################################
            # 5. Assign place_id_lists for new rows with query to database
            #####################################################################################################
            docs_structure_list = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                 docs_type).get("content")
            structure_list_for_dict = []
            for i in docs_structure_list:
                del i[1]
                del i[1]
                structure_list_for_dict.append(i)
            if docs_structure_list:
                place_id_dict = self.get_place_id_dict(structure_list_for_dict)

                self.iterator = QTreeWidgetItemIterator(tree)
                while self.iterator.value():
                    if hasattr(self.iterator.value(), 'place_id_list'):
                        self.iterator.value().place_id_list = place_id_dict[self.iterator.value().folder_name]
                    self.iterator += 1

            edit_slider.slideInIdx(0)
            self.ui.statusLabel.setText('')
            self.get_project_structure()
            get_rows_from_tree_widget()

        #####################################################################################################
        # Search rows function for tree widget
        #####################################################################################################

        def searching_patterns():
            stack_widget = None
            layout = None
            frame = None
            tree = None
            search_line = None
            search_results_label = None
            back_to_structure_btn = None
            rows_list = None
            grid = QGridLayout()
            grid.setSpacing(0)
            row_grid_index = 0
            added_items = []

            if self.ui.stackedWidget_3.currentIndex() == 0:
                stack_widget = self.ui.designDocsStructureStackedWidget
                search_line = self.ui.lineEdit_5
                search_results_label = self.ui.designDocsSearchResultsLabel
                layout = self.ui.designDocsSearchLayout
                frame = self.ui.designDocsSearchFrame
                back_to_structure_btn = self.ui.backToStructureBtn
                tree = self.ui.designDocsStructureTreeWidget
                rows_list = self.design_docs_tree_rows
                self.ui.backToStructureBtn.show()
            elif self.ui.stackedWidget_3.currentIndex() == 1:
                stack_widget = self.ui.constructionDocsStructureStackedWidget
                search_line = self.ui.lineEdit_6
                search_results_label = self.ui.constructionDocsSearchResultsLabel
                layout = self.ui.constructionDocsSearchLayout
                frame = self.ui.constructionDocsSearchFrame
                back_to_structure_btn = self.ui.backToStructureBtn_2
                tree = self.ui.constructionDocsStructureTreeWidget
                rows_list = self.construction_docs_tree_rows
                self.ui.backToStructureBtn_2.show()
            elif self.ui.stackedWidget_3.currentIndex() == 2:
                stack_widget = self.ui.initialPermitDocsStructureStackedWidget
                search_line = self.ui.lineEdit_7
                search_results_label = self.ui.initialPermitDocsSearchResultsLabel
                layout = self.ui.initialPermitDocsSearchLayout
                frame = self.ui.initialPermitDocsSearchFrame
                back_to_structure_btn = self.ui.backToStructureBtn_3
                tree = self.ui.initialPermitDocsStructureTreeWidget
                rows_list = self.initial_permit_docs_tree_rows
                self.ui.backToStructureBtn_3.show()

            for i in frame.children():
                if isinstance(i, (SearchResult,)):
                    i.setParent(None)
            for i in reversed(range(layout.count())):
                layout_item = layout.itemAt(i)
                if isinstance(layout_item, (QGridLayout,)):
                    layout_item.setParent(None)
                if isinstance(layout_item, (QSpacerItem,)):
                    layout.removeItem(layout_item)

            stack_widget.setCurrentIndex(1)

            if search_line.text() == '':
                stack_widget.setCurrentIndex(0)
                back_to_structure_btn.hide()
            elif search_line.text():
                back_to_structure_btn.show()
                for i in rows_list:
                    reference = i[1]
                    duplicate_num = 0
                    place_id_list = []
                    found = i[1].lower().find(search_line.text())
                    if found != -1:
                        for j in rows_list:
                            if j[1] == reference:
                                duplicate_num += 1
                                place_id_list.append(j[0])
                        if duplicate_num == 1:
                            row = SearchResult()
                            row.label.setText(i[1])
                            grid.addWidget(row, row_grid_index, 0, Qt.AlignTop)
                            row.place_id = place_id_list[0]
                            row.label.release.connect(lambda: back_to_structure_btn.hide())
                            row.label2.release.connect(lambda: back_to_structure_btn.hide())
                            added_items.append([row.place_id, row.label.text()])
                            row_grid_index += 1
                        else:
                            for k in range(duplicate_num):
                                if [place_id_list[k], i[1]] in added_items:
                                    pass
                                else:
                                    row = SearchResult()
                                    row.label.setText(i[1])
                                    grid.addWidget(row, row_grid_index, 0, Qt.AlignTop)
                                    row.place_id = place_id_list[k]
                                    row.label.release.connect(lambda: back_to_structure_btn.hide())
                                    row.label2.release.connect(lambda: back_to_structure_btn.hide())
                                    row.duplicate = True
                                    added_items.append([row.place_id, row.label.text()])
                                    row_grid_index += 1

                layout.addLayout(grid)
                spacer_item = QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding)
                layout.addSpacerItem(spacer_item)
                for i in frame.findChildren(SearchResult):
                    checked_rows = []
                    if i.duplicate:
                        iterator = QTreeWidgetItemIterator(tree)
                        while iterator.value():
                            item = iterator.value()
                            if isinstance(item, (RowElement,)):
                                checked_rows.append(item)
                                if item.place_id_list == i.place_id and \
                                        item.folder_name.lower() == i.label.text().lower():
                                    for j in checked_rows:
                                        if not j.parent():
                                            i.label2.setText(f'found at\n"{j.folder_name}"')
                                    break
                            iterator += 1

                if grid.itemAt(0):
                    search_results_label.setText(u'Search results:')
                else:
                    search_results_label.setText(u'Search results: nothing found')

        def load_structure_from_file():
            tree = None
            table_type = None

            self.dialog = QFileDialog()
            filename = self.dialog.getOpenFileName(None, "Select excel tabel", "", "Files (*.xlsx)")[0]
            file = load_workbook(filename, read_only=True)
            first_sheet = file.sheetnames[1]
            work_sheet = file[first_sheet]

            if self.ui.stackedWidget_3.currentIndex() == 0:
                tree = self.ui.designDocsStructureTreeWidget
                table_type = 'design'
            if self.ui.stackedWidget_3.currentIndex() == 1:
                tree = self.ui.constructionDocsStructureTreeWidget
                table_type = 'construction'
            if self.ui.stackedWidget_3.currentIndex() == 2:
                tree = self.ui.initialPermitDocsStructureTreeWidget
                table_type = 'init_permit'

            if filename:

                rows = []
                for row in work_sheet.rows:
                    rows.append(row)

                work_rows = rows[1:]
                structure_elements_to_add = {}
                docs_to_add = {}

                self.logic_index = None
                for i in range(len(work_rows)):
                    # print(work_rows[i][0].value)
                    if work_rows[i][0].value:
                        if work_rows[i][1].value:
                            text = ' '.join(work_rows[i][1].value.strip().replace('\n', ' ').split())
                            self.logic_index = (str(work_rows[i][0].value).strip()).split('.')
                            structure_elements_to_add[tuple(self.logic_index)] = f"{'.'.join(self.logic_index)}. " \
                                                                                 f"{text}"
                        else:
                            doc_data = {'doc_name': work_rows[i][2].value, 'doc_cypher': work_rows[i][3].value.strip(),
                                        'ITN': work_rows[i][4].value.strip()}
                            if tuple(self.logic_index) in docs_to_add.keys():
                                docs_to_add[tuple(self.logic_index)].append(doc_data)
                            else:
                                docs_to_add[tuple(self.logic_index)] = []
                                docs_to_add[tuple(self.logic_index)].append(doc_data)
                    else:
                        print('WRONG DATA')
                        structure_elements_to_add.clear()
                        docs_to_add.clear()
                        break

                        ################################################################
                        # there should be information window about incorrect file loaded
                        ################################################################

                added_items = []
                self.current_item = None
                self.item_to_add_to = None
                self.new_child_item = None

                if structure_elements_to_add:

                    items_to_remove = []
                    iterator = QTreeWidgetItemIterator(tree)
                    while iterator.value():
                        item = iterator.value()
                        if isinstance(item, (RowElement,)):
                            items_to_remove.append(item)
                        iterator += 1
                    for i in items_to_remove:
                        i.remove()

                    for k, v in structure_elements_to_add.items():
                        if len(k) == 1:
                            self.current_item = add_top_row(v)
                            self.current_item.logical_index = k
                            self.current_item.nesting_level = 1
                            added_items.append(self.current_item)
                        if len(k) > 1:
                            item_logical_index_to_add_to = list(k)
                            pop = item_logical_index_to_add_to.pop(-1)
                            for i in added_items:
                                if i.logical_index == tuple(item_logical_index_to_add_to):
                                    self.item_to_add_to = i
                            self.new_child_item = self.item_to_add_to.add_child()
                            self.new_child_item.frame.hide()
                            self.new_child_item.lineEdit.setReadOnly(True)
                            self.new_child_item.lineEdit.clearFocus()
                            self.new_child_item.parent().setExpanded(True)
                            self.new_child_item.lineEdit.insertPlainText(v)
                            self.new_child_item.logical_index = k
                            self.new_child_item.nesting_level = len(k)
                            added_items.append(self.new_child_item)

                    end_edit_structure()

                    collected_place_ids_and_logical_indexes = {}

                    iterator = QTreeWidgetItemIterator(tree)
                    while iterator.value():
                        item = iterator.value()
                        if isinstance(item, (RowElement,)):
                            index = []
                            for k in item.folder_name.split('.'):
                                try:
                                    intk = int(k)
                                    index.append(f'{intk}')
                                except ValueError:
                                    break
                            if index:
                                item.logical_index = tuple(index)
                                collected_place_ids_and_logical_indexes[item.logical_index] = \
                                    (tuple(item.place_id_list), item.folder_name)
                        iterator += 1

                    push_data_dict = {}

                    for k, v in docs_to_add.items():
                        new_key = collected_place_ids_and_logical_indexes.get(k)
                        new_value = v
                        push_data_dict[new_key] = new_value
                    # print(push_data_dict)

                    add_doc_process = AddDocDialog(multiple_loading_dict=push_data_dict, window_object=self)
                    add_doc_process.add_document()
                    self.get_current_project_docs_dicts_list()
                    self.fill_table(doc_type=table_type)

        def add_top_row(text=None):
            self.item = None
            for tree in self.ui.stackedWidget_3.findChildren(QTreeWidget):
                page = tree.parent().parent().parent().parent()
                current_page = self.ui.stackedWidget_3.currentWidget()
                if current_page == page:
                    if text:
                        self.item = RowElement(tree, window_object=self)
                        self.item.lineEdit.insertPlainText(text.strip())
                        self.item.lineEdit.setReadOnly(True)
                    else:
                        self.item = RowElement(tree, window_object=self)
                        self.item.lineEdit.setReadOnly(False)
                        self.item.lineEdit.setFocus()
                        self.item.frame.show()
            return self.item

        # loginPage elements

        self.ui.regBtn.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(1))
        self.ui.regBtn.clicked.connect(lambda: self.ui.emailEntering.setText(''))
        self.ui.regBtn.clicked.connect(lambda: self.ui.passEntering.setText(''))
        self.ui.restorePassBtn.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(6))
        self.ui.restorePassBtn.clicked.connect(lambda: self.ui.emailEntering.setText(''))
        self.ui.restorePassBtn.clicked.connect(lambda: self.ui.passEntering.setText(''))
        self.ui.loginBtn.clicked.connect(lambda: login_with_saving_option())

        # regPage1 buttons

        self.ui.regBtn_2.clicked.connect(lambda: sending_key())
        self.ui.cancelRegBtn.clicked.connect(lambda: new_user_data_clear())
        self.ui.cancelRegBtn.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # regPage2 buttons
        self.ui.proceedBtn.clicked.connect(lambda: key_accepting())
        self.ui.cancelRegBtn_2.clicked.connect(lambda: new_user_data_clear())
        self.ui.cancelRegBtn_2.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # regPage3 buttons
        self.ui.proceedBtn_2.clicked.connect(lambda: new_user_password_entering())
        self.ui.cancelRegBtn_3.clicked.connect(lambda: new_user_data_clear())
        self.ui.cancelRegBtn_3.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # regPage4 buttons
        self.ui.nameEntering.textChanged.connect(lambda: self.ui.nameEntering.setStyleSheet(
            u'background-color: rgb(184, 184, 184); color : black;'))
        self.ui.lastNameEntering.textChanged.connect(lambda: self.ui.lastNameEntering.setStyleSheet(
            u'background-color: rgb(184, 184, 184); color : black;'))
        self.ui.companyTIN.textChanged.connect(lambda: self.ui.companyTIN.setStyleSheet(
            u'background-color: rgb(184, 184, 184); color : black;'))
        self.ui.proceedBtn_3.clicked.connect(lambda: create_user())
        self.ui.cancelRegBtn_4.clicked.connect(lambda: new_user_data_clear())
        self.ui.cancelRegBtn_4.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # regPage5 buttons
        self.ui.goLoginBtn.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # regPage6 buttons
        self.ui.restorePasswordBtn.clicked.connect(lambda: sending_recover_key())
        self.ui.cancelRegBtn_5.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))
        self.ui.cancelRegBtn_5.clicked.connect(lambda: exist_user_data_clear())

        # regPage7 buttons
        self.ui.proceedBtn_4.clicked.connect(lambda: key_recover_accepting())
        self.ui.cancelRegBtn_6.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))
        self.ui.cancelRegBtn_6.clicked.connect(lambda: exist_user_data_clear())

        # regPage8 buttons
        self.ui.proceedBtn_5.clicked.connect(lambda: user_password_recover())
        self.ui.cancelRegBtn_7.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))
        self.ui.cancelRegBtn_7.clicked.connect(lambda: exist_user_data_clear())

        # regPage9 buttons
        self.ui.goLoginBtn_2.clicked.connect(lambda: self.ui.regStackedWidget.slideInIdx(0))

        # main_interface_sub_menu_buttons
        self.ui.logOutBtn.clicked.connect(lambda: self.close_app(log_out=True))
        self.ui.logOutBtn.clicked.connect(lambda: self.session.api.logout)
        self.ui.logOutBtn.clicked.connect(lambda: self.ui.mainMenuStack.slideInIdx(0))
        self.ui.logOutBtn.clicked.connect(lambda: clear_widgets())
        self.ui.logOutBtn.clicked.connect(lambda: self.ui.statusLabel.setText(''))
        self.ui.logOutBtn.clicked.connect(self.clear_projects)
        self.ui.loginBtn.clicked.connect(lambda: self.ui.rememberCheckBox.setChecked(False))
        self.ui.notificationsBtn.clicked.connect(lambda: self.ui.NotificationsMenu.hide_show_func())

        # new project creation
        self.ui.newProjectBtn.clicked.connect(lambda: self.ui.interfaceBodyStackedWidget.slideInIdx(0))
        self.ui.newProjectBtn.clicked.connect(lambda: self.ui.newProjectStackedWidget.slideInIdx(0))
        self.ui.lineEdit_4.clicked.connect(lambda: get_picture_filepath())
        self.ui.proceedNewProject.clicked.connect(lambda: new_project_adding_process_page1())
        self.ui.backToNewProjectMenuBtn.clicked.connect(
            lambda: self.ui.newProjectStackedWidget.slideInIdx(0))
        self.ui.backToNewProjectMenuBtn.clicked.connect(lambda: new_project_data_end_clear())
        self.ui.cancelNewProjectBtn.clicked.connect(
            lambda: self.ui.interfaceBodyStackedWidget.slideInIdx(1))
        self.ui.lineEdit_4.textChanged[str].connect(lambda: self.ui.lineEdit_4.setStyleSheet(u"background-color: rgb("
                                                                                             u"184, 184, 184); color "
                                                                                             u": black"))
        self.ui.lineEdit_3.textChanged[str].connect(lambda: self.ui.lineEdit_3.setStyleSheet(u"background-color: rgb("
                                                                                             u"184, 184, 184); color "
                                                                                             u": black"))
        self.ui.lineEdit_2.textChanged[str].connect(lambda: self.ui.lineEdit_2.setStyleSheet(u"background-color: rgb("
                                                                                             u"184, 184, 184); color "
                                                                                             u": black"))
        self.ui.lineEdit.textChanged[str].connect(lambda: self.ui.lineEdit.setStyleSheet(u"background-color: rgb("
                                                                                         u"184, 184, 184); color "
                                                                                         u": black"))
        self.ui.lineEdit_4.textChanged[str].connect(lambda: self.ui.label_10.setText(''))
        self.ui.lineEdit_3.textChanged[str].connect(lambda: self.ui.label_10.setText(''))
        self.ui.lineEdit_2.textChanged[str].connect(lambda: self.ui.label_10.setText(''))
        self.ui.lineEdit.textChanged[str].connect(lambda: self.ui.label_10.setText(''))
        self.ui.createNewProject.clicked.connect(lambda: new_project_adding_process_page2())
        self.ui.cancelNewProjectBtn.clicked.connect(lambda: new_project_data_begin_clear())
        self.ui.comboBox_4.currentIndexChanged.connect(lambda: set_company_users_list(self.ui.comboBox_4.currentText(),
                                                                                      self.ui.comboBox_7))
        self.ui.comboBox_5.currentIndexChanged.connect(lambda: set_company_users_list(self.ui.comboBox_5.currentText(),
                                                                                      self.ui.comboBox_8))
        self.ui.comboBox_6.currentIndexChanged.connect(lambda: set_company_users_list(self.ui.comboBox_6.currentText(),
                                                                                      self.ui.comboBox_9))
        self.ui.comboBox_3.currentIndexChanged.connect(lambda: set_company_users_list(self.ui.comboBox_3.currentText(),
                                                                                      self.ui.comboBox_2))

        # tree view elements

        def show_or_hide_left_side_menu_for_trigger():
            if self.ui.leftSidePopUpMenu.isHidden():
                self.ui.leftSidePopUpMenu.show_animate()
            else:
                self.ui.leftSidePopUpMenu.hide_animate()

        def show_or_hide_left_side_menu_for_home_btn():
            if not self.ui.leftSidePopUpMenu.isHidden():
                self.ui.leftSidePopUpMenu.hide()

        self.ui.homeBtn.clicked.connect(lambda: self.ui.homeBtn.hide())
        self.ui.homeBtn.clicked.connect(lambda: self.ui.leftSideMenuBtn.hide())
        self.ui.leftSideMenuBtn.clicked.connect(lambda: show_or_hide_left_side_menu_for_trigger())
        self.ui.homeBtn.clicked.connect(lambda: self.ui.newProjectBtn.show())
        self.ui.homeBtn.clicked.connect(lambda: self.ui.editProjectCardBtn.show())
        self.ui.homeBtn.clicked.connect(self.clear_tables)
        self.ui.homeBtn.clicked.connect(lambda: show_or_hide_left_side_menu_for_home_btn())
        self.ui.editStructureBtn.clicked.connect(lambda: self.ui.stackedWidget_4.slideInIdx(1))
        self.ui.editStructureBtn.clicked.connect(lambda: edit_structure())
        self.ui.editStructureBtn_2.clicked.connect(lambda: self.ui.stackedWidget_5.slideInIdx(1))
        self.ui.editStructureBtn_2.clicked.connect(lambda: edit_structure())
        self.ui.editStructureBtn_3.clicked.connect(lambda: self.ui.stackedWidget_6.slideInIdx(1))
        self.ui.editStructureBtn_3.clicked.connect(lambda: edit_structure())
        self.ui.addSubContainerBtn.clicked.connect(lambda: add_top_row())
        self.ui.addSubContainerBtn_2.clicked.connect(lambda: add_top_row())
        self.ui.addSubContainerBtn_3.clicked.connect(lambda: add_top_row())
        self.ui.endEditBtn.clicked.connect(lambda: end_edit_structure())
        self.ui.endEditBtn_2.clicked.connect(lambda: end_edit_structure())
        self.ui.endEditBtn_3.clicked.connect(lambda: end_edit_structure())
        self.ui.cancelEditBtn.clicked.connect(lambda: self.ui.stackedWidget_4.slideInIdx(0))
        self.ui.cancelEditBtn.clicked.connect(lambda: cancel_edit_structure())
        self.ui.cancelEditBtn_2.clicked.connect(lambda: cancel_edit_structure())
        self.ui.cancelEditBtn_2.clicked.connect(lambda: self.ui.stackedWidget_5.slideInIdx(0))
        self.ui.cancelEditBtn_3.clicked.connect(lambda: cancel_edit_structure())
        self.ui.cancelEditBtn_3.clicked.connect(lambda: self.ui.stackedWidget_6.slideInIdx(0))
        self.ui.lineEdit_5.textEdited.connect(lambda: searching_patterns())
        self.ui.lineEdit_6.textEdited.connect(lambda: searching_patterns())
        self.ui.lineEdit_7.textEdited.connect(lambda: searching_patterns())
        self.ui.lineEdit_5.clicked.connect(lambda: get_rows_from_tree_widget())
        self.ui.lineEdit_6.clicked.connect(lambda: get_rows_from_tree_widget())
        self.ui.lineEdit_7.clicked.connect(lambda: get_rows_from_tree_widget())
        self.ui.backToStructureBtn.clicked.connect(lambda: self.ui.designDocsStructureStackedWidget.setCurrentIndex(0))
        self.ui.backToStructureBtn_2.clicked.connect(
            lambda: self.ui.constructionDocsStructureStackedWidget.setCurrentIndex(0))
        self.ui.backToStructureBtn_3.clicked.connect(
            lambda: self.ui.initialPermitDocsStructureStackedWidget.setCurrentIndex(0))
        self.ui.backToStructureBtn.clicked.connect(lambda: self.ui.backToStructureBtn.hide())
        self.ui.backToStructureBtn_2.clicked.connect(lambda: self.ui.backToStructureBtn_2.hide())
        self.ui.backToStructureBtn_3.clicked.connect(lambda: self.ui.backToStructureBtn_3.hide())

        self.ui.searchBarBtn.clicked.connect(lambda: self.ui.frame_26.hide_show_func())
        self.ui.searchBarBtn_2.clicked.connect(lambda: self.ui.frame_32.hide_show_func())
        self.ui.searchBarBtn_3.clicked.connect(lambda: self.ui.frame_36.hide_show_func())

        self.ui.backToStructureBtn.hide()
        self.ui.lineEdit_5.clearFocus()
        self.ui.backToStructureBtn_2.hide()
        self.ui.lineEdit_6.clearFocus()
        self.ui.backToStructureBtn_3.hide()
        self.ui.lineEdit_7.clearFocus()

        def structure_search_line_hiding():
            self.doc_search_line = None
            if self.ui.stackedWidget_3.currentIndex() == 0:
                self.doc_search_line = self.ui.lineEdit_5
            if self.ui.stackedWidget_3.currentIndex() == 1:
                self.doc_search_line = self.ui.lineEdit_6
            if self.ui.stackedWidget_3.currentIndex() == 2:
                self.doc_search_line = self.ui.lineEdit_7
            if self.doc_search_line.isHidden():
                self.doc_search_line.show()
            else:
                self.doc_search_line.hide()

        def structure_release_btn_showing():
            self.doc_search_line = None
            self.release_button = None
            if self.ui.stackedWidget_3.currentIndex() == 0:
                self.doc_search_line = self.ui.lineEdit_5
                self.release_button = self.ui.backToStructureBtn
            if self.ui.tabWidget.currentIndex() == 1:
                self.doc_search_line = self.ui.lineEdit_6
                self.release_button = self.ui.backToStructureBtn_2
            if self.ui.tabWidget.currentIndex() == 2:
                self.doc_search_line = self.ui.lineEdit_7
                self.release_button = self.ui.backToStructureBtn_3
            if self.doc_search_line.text() == '':
                self.release_button.hide()
            else:
                self.release_button.show()

        self.ui.searchBarBtn.clicked.connect(lambda: structure_search_line_hiding())
        self.ui.searchBarBtn_2.clicked.connect(lambda: structure_search_line_hiding())
        self.ui.searchBarBtn_3.clicked.connect(lambda: structure_search_line_hiding())
        self.ui.lineEdit_5.hide()
        self.ui.lineEdit_6.hide()
        self.ui.lineEdit_7.hide()

        self.ui.lineEdit_5.textEdited.connect(lambda: structure_release_btn_showing())
        self.ui.lineEdit_6.textEdited.connect(lambda: structure_release_btn_showing())
        self.ui.lineEdit_7.textEdited.connect(lambda: structure_release_btn_showing())

        self.ui.backToStructureBtn.clicked.connect(lambda: self.ui.lineEdit_5.setText(''))
        self.ui.backToStructureBtn.clicked.connect(lambda: self.ui.backToStructureBtn.hide())
        self.ui.backToStructureBtn_2.clicked.connect(lambda: self.ui.lineEdit_6.setText(''))
        self.ui.backToStructureBtn_2.clicked.connect(lambda: self.ui.backToStructureBtn_2.hide())
        self.ui.backToStructureBtn_3.clicked.connect(lambda: self.ui.lineEdit_7.setText(''))
        self.ui.backToStructureBtn_3.clicked.connect(lambda: self.ui.backToStructureBtn_3.hide())
        self.ui.searchBarBtn.clicked.connect(lambda: self.ui.backToStructureBtn.hide())
        self.ui.searchBarBtn_2.clicked.connect(lambda: self.ui.backToStructureBtn_2.hide())
        self.ui.searchBarBtn_3.clicked.connect(lambda: self.ui.backToStructureBtn_3.hide())

        self.ui.frame_26.expanded_width = 200
        self.ui.frame_32.expanded_width = 200
        self.ui.frame_36.expanded_width = 200
        self.ui.frame_26.setFixedWidth(30)
        self.ui.frame_32.setFixedWidth(30)
        self.ui.frame_36.setFixedWidth(30)

        # table view elements

        self.ui.searchDocsBtn.clicked.connect(lambda: self.ui.searchDocsFrame.hide_show_func())
        self.ui.searchDocsBtn_2.clicked.connect(lambda: self.ui.searchDocsFrame_2.hide_show_func())
        self.ui.searchDocsBtn_3.clicked.connect(lambda: self.ui.searchDocsFrame_3.hide_show_func())
        self.ui.searchDocsLine.textChanged.connect(lambda: self.fill_table(doc_type='design',
                                                                           text_to_search=self.ui.searchDocsLine.text()))
        self.ui.searchDocsLine_2.textChanged.connect(lambda: self.fill_table(doc_type='construction',
                                                                             text_to_search=self.ui.searchDocsLine_2.text()))
        self.ui.searchDocsLine_3.textChanged.connect(lambda: self.fill_table(doc_type='construction',
                                                                             text_to_search=self.ui.searchDocsLine_3.text()))

        self.ui.releaseBtn.hide()
        self.ui.searchDocsLine.clearFocus()
        self.ui.releaseBtn_2.hide()
        self.ui.searchDocsLine_2.clearFocus()
        self.ui.releaseBtn_3.hide()
        self.ui.searchDocsLine_3.clearFocus()

        def search_line_hiding():
            self.doc_search_line = None
            if self.ui.tabWidget.currentIndex() == 0:
                self.doc_search_line = self.ui.searchDocsLine
            if self.ui.tabWidget.currentIndex() == 1:
                self.doc_search_line = self.ui.searchDocsLine_2
            if self.ui.tabWidget.currentIndex() == 2:
                self.doc_search_line = self.ui.searchDocsLine_3
            if self.doc_search_line.isHidden():
                self.doc_search_line.show()
            else:
                self.doc_search_line.hide()

        def release_btn_showing():
            self.doc_search_line = None
            self.release_button = None
            if self.ui.tabWidget.currentIndex() == 0:
                self.doc_search_line = self.ui.searchDocsLine
                self.release_button = self.ui.releaseBtn
            if self.ui.tabWidget.currentIndex() == 1:
                self.doc_search_line = self.ui.searchDocsLine_2
                self.release_button = self.ui.releaseBtn_2
            if self.ui.tabWidget.currentIndex() == 2:
                self.doc_search_line = self.ui.searchDocsLine_3
                self.release_button = self.ui.releaseBtn_3
            if self.doc_search_line.text() == '':
                self.release_button.hide()
            else:
                self.release_button.show()

        self.ui.searchDocsBtn.clicked.connect(lambda: search_line_hiding())
        self.ui.searchDocsBtn_2.clicked.connect(lambda: search_line_hiding())
        self.ui.searchDocsBtn_3.clicked.connect(lambda: search_line_hiding())
        self.ui.searchDocsLine.hide()
        self.ui.searchDocsLine_2.hide()
        self.ui.searchDocsLine_3.hide()

        self.import_from_excel = QAction('Import from Excel')
        self.export_to_excel = QAction('Export to Excel')
        self.download_template = QAction('Download Template')

        self.context_menu = QMenu()

        self.context_menu.addAction(self.import_from_excel)
        self.context_menu.addAction(self.export_to_excel)
        self.context_menu.addAction(self.download_template)

        self.import_from_excel.triggered.connect(lambda: load_structure_from_file())
        # self.export_to_excel.triggered.connect(lambda: )
        self.download_template.triggered.connect(lambda: self.session.download_template())

        self.ui.subMenuBtn.setMenu(self.context_menu)
        self.ui.subMenuBtn_2.setMenu(self.context_menu)
        self.ui.subMenuBtn_3.setMenu(self.context_menu)
        self.ui.searchDocsLine.textEdited.connect(lambda: release_btn_showing())
        self.ui.searchDocsLine_2.textEdited.connect(lambda: release_btn_showing())
        self.ui.searchDocsLine_3.textEdited.connect(lambda: release_btn_showing())
        self.ui.releaseBtn.clicked.connect(lambda: self.ui.searchDocsLine.setText(''))
        self.ui.releaseBtn.clicked.connect(lambda: self.ui.releaseBtn.hide())
        self.ui.releaseBtn_2.clicked.connect(lambda: self.ui.searchDocsLine_2.setText(''))
        self.ui.releaseBtn_2.clicked.connect(lambda: self.ui.releaseBtn_2.hide())
        self.ui.releaseBtn_3.clicked.connect(lambda: self.ui.searchDocsLine_3.setText(''))
        self.ui.releaseBtn_3.clicked.connect(lambda: self.ui.releaseBtn_3.hide())
        self.ui.searchDocsBtn.clicked.connect(lambda: self.ui.releaseBtn.hide())
        self.ui.searchDocsBtn_2.clicked.connect(lambda: self.ui.releaseBtn_2.hide())
        self.ui.searchDocsBtn_3.clicked.connect(lambda: self.ui.releaseBtn_3.hide())
        self.ui.addDocumentBtn.clicked.connect(lambda: self.add_document_dialog())
        self.ui.addDocumentBtn_2.clicked.connect(lambda: self.add_document_dialog())
        self.ui.addDocumentBtn_3.clicked.connect(lambda: self.add_document_dialog())

        self.header_design_docs_table = Header(
            columns_names_list=['Part', 'Cypher', 'Name', 'Author', 'Version', 'Status', 'Exam result',
                                'Rev after exam', 'Start Develop', 'End Develop'],
            parent=self.ui.designDocsTableWidget)
        self.header_construction_docs_table = Header(
            columns_names_list=['Part', 'Cypher', 'Name', 'Author', 'Version', 'Status', 'Exam result',
                                'Rev after exam', 'Start Develop', 'End Develop'],
            parent=self.ui.constructionDocsTableWidget)
        self.header_init_permit_docs_table = Header(
            columns_names_list=['Part', 'Cypher', 'Name', 'Author', 'Version', 'Status', 'Exam result',
                                'Rev after exam', 'Start Develop', 'End Develop'],
            parent=self.ui.initDocsTableWidget)

        self.ui.designDocsTableWidget.setColumnWidth(0, 150)
        self.ui.designDocsTableWidget.setColumnWidth(1, 180)
        self.ui.designDocsTableWidget.setColumnWidth(2, 200)
        self.ui.designDocsTableWidget.setColumnWidth(3, 150)
        self.ui.designDocsTableWidget.setColumnWidth(4, 80)
        self.ui.designDocsTableWidget.setColumnWidth(5, 150)

        self.ui.designDocsTableWidget.verticalHeader().setVisible(False)
        self.ui.constructionDocsTableWidget.verticalHeader().setVisible(False)
        self.ui.initDocsTableWidget.verticalHeader().setVisible(False)
        self.ui.designDocsTableWidget.setHorizontalHeaderLabels('        ')
        self.ui.constructionDocsTableWidget.setHorizontalHeaderLabels('        ')
        self.ui.initDocsTableWidget.setHorizontalHeaderLabels('        ')

        self.ui.mainHeader.closeBtn.clicked.connect(lambda: self.close_app())

    def add_child(self, parent, name, place_id_list):
        child = RowElement(parent, window_object=self)
        child.lineEdit.setTextInteractionFlags(Qt.TextInteractionFlag.NoTextInteraction)
        child.lineEdit.setTextInteractionFlags(Qt.TextInteractionFlag.TextEditorInteraction)
        child.lineEdit.setTextInteractionFlags(Qt.TextInteractionFlag.TextEditable)
        child.lineEdit.setCursorWidth(0)
        child.lineEdit.insertPlainText(name.strip())
        child.folder_name = name
        child.previous_folder_name = name
        child.place_id_list = place_id_list
        child.lineEdit.viewport().setProperty("cursor", QCursor(Qt.ArrowCursor))
        child.setExpanded(True)

    def get_place_id_dict(self, docs_structure_list):
        if docs_structure_list:
            place_id_dict = {}
            for i in docs_structure_list:
                for j in i[1:]:
                    if j:
                        place_id_dict[j] = []
            for i in docs_structure_list:
                for k in place_id_dict:
                    if k in i:
                        place_id_dict[k].append(i[0])
            return place_id_dict

    def sort_structure_list(self, list_to_sort):

        sorted_list = []
        elements_dict = {}

        for i in range(len(list_to_sort)):
            visual_num = []
            for k in list_to_sort[i][1].split('.')[0:2]:
                try:
                    intk = int(k)
                    visual_num.append(intk)
                except ValueError:
                    break
            if len(visual_num):
                if len(visual_num) > 1:
                    elements_dict[i] = sum(visual_num)
                elements_dict[i] = visual_num
            else:
                elements_dict[i] = []

        sorted_indexes_dict = sorted(elements_dict.items(), key=lambda x: x[1])

        items_to_take = []
        for j in range(len(sorted_indexes_dict)):
            if not sorted_indexes_dict[j][1]:
                items_to_take.append(sorted_indexes_dict[j])
        for i in items_to_take:
            sorted_indexes_dict.remove(i)
            sorted_indexes_dict.append(i)

        for i in sorted_indexes_dict:
            sorted_list.append(list_to_sort[i[0]])

        return sorted_list

    def apply_structure(self):

        for index in range(self.ui.stackedWidget_3.count()):
            structure_list = None
            tree = None
            if index == 0:
                structure_list = self.sort_structure_list(self.design_docs_structure_list)
                tree = self.ui.designDocsStructureTreeWidget
            elif index == 1:
                structure_list = self.sort_structure_list(self.construction_docs_structure_list)
                tree = self.ui.constructionDocsStructureTreeWidget
            elif index == 2:
                structure_list = self.sort_structure_list(self.initial_permit_docs_structure_list)
                tree = self.ui.initialPermitDocsStructureTreeWidget

            if structure_list:

                ###########################################################################################
                # Make dict to assign list of 'place_id' values to each folder_name
                # this is necessary to have ability to sorting documents if some folder selected
                ###########################################################################################
                place_id_dict = self.get_place_id_dict(structure_list)

                self.prev_i = None
                added_rows = []
                for i in structure_list:
                    if i[1] not in added_rows:
                        self.parent = RowElement(tree, window_object=self)

                        doc = self.parent.lineEdit.document()
                        cursor = QTextCursor(doc)
                        cursor.insertText(i[1])
                        self.parent.previous_folder_name = i[1]

                        self.parent.place_id_list = place_id_dict[i[1]]
                        self.parent.folder_name = i[1]
                        self.parent.lineEdit.viewport().setProperty("cursor", QCursor(Qt.ArrowCursor))
                        self.parent.setExpanded(True)
                    added_rows.append(i[1])

                self.iterator = QTreeWidgetItemIterator(tree)
                top_level_elements_count = 0

                while self.iterator.value():
                    top_level_elements_count += 1
                    self.iterator += 1

                index = 0
                for i in structure_list:
                    if len(i) > index:
                        index = len(i)

                for top_element_index in range(top_level_elements_count)[1:]:
                    for ind in range(index - 1):
                        self.item = tree.topLevelItem(top_element_index)
                        self.iterator2 = QTreeWidgetItemIterator(self.item, QTreeWidgetItemIterator.NoChildren)
                        self.parent = None
                        while self.iterator2.value():
                            self.parent = self.iterator2.value()
                            self.pair = []
                            self.prev_child_num = 0
                            for i in structure_list:
                                if i[ind + 1]:
                                    self.pair = [i[ind], i[ind + 1]]
                                    if self.pair[0] == self.parent.lineEdit.document().toPlainText():
                                        if hasattr(self.parent.child(self.prev_child_num - 1), 'lineEdit'):
                                            if self.parent.child(
                                                    self.prev_child_num - 1).lineEdit.document().toPlainText() != \
                                                    self.pair[1]:
                                                place_id_list = []
                                                for place_id in place_id_dict[self.pair[1]]:
                                                    if place_id in self.parent.place_id_list:
                                                        place_id_list.append(place_id)
                                                if place_id_list:
                                                    iterator = QTreeWidgetItemIterator(self.parent)
                                                    added_children = []
                                                    while iterator.value():
                                                        item = iterator.value()
                                                        added_children.append(
                                                            (item.place_id_list, item.folder_name))
                                                        iterator += 1
                                                    if (place_id_list, self.pair[1]) not in added_children:
                                                        self.add_child(self.parent, self.pair[1], place_id_list)
                                                        self.prev_child_num = self.parent.childCount()
                                        else:
                                            place_id_list = []
                                            for place_id in place_id_dict[self.pair[1]]:
                                                if place_id in self.parent.place_id_list:
                                                    place_id_list.append(place_id)
                                            if place_id_list:
                                                iterator = QTreeWidgetItemIterator(self.parent)
                                                added_children = []
                                                while iterator.value():
                                                    item = iterator.value()
                                                    added_children.append(
                                                        (item.place_id_list, item.folder_name))
                                                    iterator += 1
                                                if (place_id_list, self.pair[1]) not in added_children:
                                                    self.add_child(self.parent, self.pair[1], place_id_list)
                                                    self.prev_child_num = self.parent.childCount()
                            self.iterator2 += 1

    def get_project_structure(self):
        self.ui.designDocsStructureTreeWidget.clear()
        self.ui.constructionDocsStructureTreeWidget.clear()
        self.ui.initialPermitDocsStructureTreeWidget.clear()
        RowElement2(self.ui.designDocsStructureTreeWidget, window_object=self)
        RowElement2(self.ui.constructionDocsStructureTreeWidget, window_object=self)
        RowElement2(self.ui.initialPermitDocsStructureTreeWidget, window_object=self)

        if self.current_project_id:
            design_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                      'design').get("content")
            self.design_docs_structure_list = []
            if design_docs_query_result:
                for i in design_docs_query_result:
                    list_element = list(i)
                    del list_element[1]
                    del list_element[1]
                    self.design_docs_structure_list.append(list_element)

            construction_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                            'construction').get("content")
            self.construction_docs_structure_list = []
            if construction_docs_query_result:
                for i in construction_docs_query_result:
                    list_element = list(i)
                    del list_element[1]
                    del list_element[1]
                    self.construction_docs_structure_list.append(list_element)

            initial_permit_docs_query_result = self.session.api.get_structure(self.current_project_data_dict['id'],
                                                                              'init_permit').get("content")
            self.initial_permit_docs_structure_list = []
            if initial_permit_docs_query_result:
                for i in initial_permit_docs_query_result:
                    list_element = list(i)
                    del list_element[1]
                    del list_element[1]
                    self.initial_permit_docs_structure_list.append(list_element)

        self.apply_structure()

    def get_current_project_docs_dicts_list(self):
        response = self.session.api.get_project_docs(self.current_project_data_dict['id'])
        if response.get("content"):
            self.current_project_docs_dicts_list = response.get("content")

    def clear_tables(self):
        self.current_project_id = None
        self.ui.folderNameLabel.setText('All Documents')
        self.current_design_docs_folder = None
        self.current_design_docs_folder_path = 'All Documents'
        self.current_construction_docs_folder = None
        self.current_construction_docs_folder_path = 'All Documents'
        self.current_init_permission_docs_folder = None
        self.current_init_permission_docs_folder_path = 'All Documents'
        tables_list = [self.ui.designDocsTableWidget, self.ui.constructionDocsTableWidget, self.ui.initDocsTableWidget]
        for i in tables_list:
            i.clear()
            if i.has_left_pinned:
                for cell in i.left_pinned_table.horizontalHeader().cells:
                    cell.unpin_column(cell.true_logical_index, cell.pinned_column_logical_index)
            if i.has_right_pinned:
                for cell in i.right_pinned_table.horizontalHeader().cells:
                    cell.unpin_column(cell.true_logical_index, cell.pinned_column_logical_index)

    def clear_projects(self):
        self.project_widget_list = []

    def fill_table(self, doc_type: str = None, search_mode: bool = None, text_to_search: str = None):
        table = None
        current_folder = None
        folder_name_label = None
        current_folder_var = None
        if doc_type == 'design':
            table = self.ui.designDocsTableWidget
            current_folder = self.current_design_docs_folder
            folder_name_label = self.ui.folderNameLabel
            current_folder_var = self.current_design_docs_folder
        if doc_type == 'construction':
            table = self.ui.constructionDocsTableWidget
            current_folder = self.current_construction_docs_folder
            folder_name_label = self.ui.folderNameLabel_2
            current_folder_var = self.current_construction_docs_folder
        if doc_type == 'init_permit':
            table = self.ui.initDocsTableWidget
            current_folder = self.current_init_permission_docs_folder
            folder_name_label = self.ui.folderNameLabel_3
            current_folder_var = self.current_init_permission_docs_folder

        if self.current_project_docs_dicts_list:
            row_num = 0
            docs_to_show = []
            #######################################################################
            # SET ROWS NUMBER FOR TABLE
            for doc_dict in self.current_project_docs_dicts_list:
                if doc_dict['document_type'] == doc_type:
                    if not text_to_search:
                        if not current_folder:
                            folder_name_label.setText('All documents')
                            row_num += 1
                            docs_to_show.append(doc_dict)
                        else:
                            folder_name_label.setText(current_folder_var[1])
                            if doc_dict['place_id'] != 'None':
                                if set(ast.literal_eval(doc_dict['place_id'])[0]).issubset(set(current_folder[0])):
                                    row_num += 1
                                    docs_to_show.append(doc_dict)
                    else:
                        folder_name_label.setText('Search results')
                        if doc_dict['document_name'].find(text_to_search) >= 0 or \
                                doc_dict['document_cypher'].find(text_to_search) >= 0:
                            docs_to_show.append(doc_dict)
                            row_num += 1
            table.setRowCount(row_num)
            for row in range(table.rowCount()):
                table.setRowHeight(row, 40)
            #######################################################################
            # FILL ROWS IN MAIN TABLE
            row = 0
            for doc in docs_to_show:
                if table == self.ui.designDocsTableWidget:
                    place_id_list = None
                    if doc['place_id'] != 'None':
                        place = ast.literal_eval(doc['place_id'])[-1]
                        place_id_list = ast.literal_eval(doc['place_id'])[0:-1]
                        item = QTableWidgetItem(place)
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 0, item)
                    else:
                        item = QTableWidgetItem('All documents')
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 0, item)

                    item = QTableWidgetItem(doc['document_cypher'])
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 1, item)

                    item = QTableWidgetItem(doc['document_name'])
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 2, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 3, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 4, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 5, item)

                    item = QTableWidgetItem('No result exam yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 6, item)

                    item = QTableWidgetItem('No result exam yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 7, item)

                    if doc['start_develop_date']:
                        item = QTableWidgetItem(
                            f"{time.strptime(doc['start_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_mday}-"
                            f"{time.strptime(doc['start_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_mon}-"
                            f"{time.strptime(doc['start_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_year}")
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 8, item)
                    else:
                        item = QTableWidgetItem('Not set')
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 8, item)

                    if doc['end_develop_date']:
                        item = QTableWidgetItem(
                            f"{time.strptime(doc['end_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_mday}-"
                            f"{time.strptime(doc['end_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_mon}-"
                            f"{time.strptime(doc['end_develop_date'], '%a, %d %b %Y %H:%M:%S %Z').tm_year}")
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 9, item)
                    else:
                        item = QTableWidgetItem('Not set')
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                                doc['document_type']))
                        table.setItem(row, 9, item)

                    row += 1

                if table == self.ui.constructionDocsTableWidget:
                    place_id_list = None
                    if doc['place_id'] != 'None':
                        place = ast.literal_eval(doc['place_id'])[-1]
                        place_id_list = ast.literal_eval(doc['place_id'])[0:-1]
                        item = QTableWidgetItem(place)
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher']))
                        table.setItem(row, 0, item)
                    else:
                        item = QTableWidgetItem('All documents')
                        item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher']))
                        table.setItem(row, 0, item)

                    item = QTableWidgetItem(doc['document_cypher'])
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 1, item)

                    item = QTableWidgetItem(doc['document_name'])
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 2, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 3, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 4, item)

                    item = QTableWidgetItem('No file yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 5, item)

                    item = QTableWidgetItem('No result exam yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 6, item)

                    item = QTableWidgetItem('No result exam yet')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 7, item)

                    item = QTableWidgetItem('start_develop_date')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 8, item)

                    item = QTableWidgetItem('end_develop_date')
                    item.setData(Qt.ItemDataRole.UserRole, (doc['doc_id'], place_id_list, doc['document_cypher'],
                                                            doc['document_type']))
                    table.setItem(row, 9, item)

                    row += 1

                if table == self.ui.initDocsTableWidget:
                    pass

            def prepare_table(table_object: QCustomTableWidget):
                table_object.horizontalHeader().show_hide_scrollbars()
                table_object.get_region()
                # table_object.set_mask(corners='LR')

            if self.ui.interfaceBodyStackedWidget.anim_group:
                self.ui.interfaceBodyStackedWidget.anim_group.finished.connect(lambda: prepare_table(table))
            else:
                prepare_table(table)

            #######################################################################
            # FILL ROWS IN SIDE TABLES IF SUCH TABLES EXISTS
            for pinned_table in (table.left_pinned_table, table.right_pinned_table):
                if pinned_table:
                    for i in table.horizontalHeader().children():
                        if isinstance(i, (HeaderCell,)):
                            for j in pinned_table.horizontalHeader().children():
                                if isinstance(j, (HeaderCell,)):
                                    if i.label.text() == j.label.text():
                                        iterate_row(table, i.logical_index, pinned_table, j.true_logical_index)
            #######################################################################

    def add_document_dialog(self):
        dlg = AddDocDialog(self, window_object=self)
        dlg.exec()

    def document_view_dialog(self, doc_id):
        dlg = DocViewDialog(main_window=self, project_id=self.current_project_data_dict['id'], doc_id=doc_id)
        dlg.exec()

    def add_notification(self, ntfcn_dict=None):
        self.ui.NotificationsMenu.insert_notification(ntfcn_dict=ntfcn_dict)

    def close_app(self, log_out=False):
        if self.session:
            self.session.close_session()
            if self.message_receiver:
                self.message_receiver.stop_broker_loop()
            if self.message_receiving_thread:
                self.message_receiving_thread.terminate()
            self.ui.NotificationsMenu.clear_notifications()
            self.ui.NotificationsMenu.hide()
        if not log_out:
            self.stop = True


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
